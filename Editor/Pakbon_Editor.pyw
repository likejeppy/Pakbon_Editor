# © 2025 Jeffrey van Houten
# Alle rechten voorbehouden.
# Dit programma is auteursrechtelijk beschermd. Gebruik, kopiëren, verspreiden of wijzigen is alleen toegestaan met uitdrukkelijke schriftelijke toestemming van de rechthebbende.
current_version = "0.3.1"
import sys
import logging
import subprocess
import os

# Configure logging
logging.basicConfig(
    filename="app.log",  # Log file name
    level=logging.DEBUG,  # Logging level (DEBUG, INFO, WARNING, ERROR, CRITICAL)
    format="%(asctime)s - %(levelname)s - %(message)s",  # Log format
    datefmt="%d-%m-%Y %H:%M:%S",  # Date format
)

logging.info(
    "*******************************************************************************************************")  # Initial log entry
logging.info("Loading application.")

def exception_handler(exc_type, exc_value, exc_traceback):
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return
    logging.error("Uncaught exception", exc_info=(exc_type, exc_value, exc_traceback))

sys.excepthook = exception_handler

def install_requirements():
    logging.info("Performing function 'install_requirements'.")

    # Check if requirements.txt exists or is empty
    if not os.path.exists("requirements.txt") or os.stat("requirements.txt").st_size == 0:
        logging.info("requirements.txt is missing or empty, continuing without installing additional libraries.")
        return  # Continue the program without exiting the program, but will exit the function

    # Read the requirements.txt file
    with open("requirements.txt", "r") as f:
        libraries = f.readlines()

    raised_error = False  # Initialize the raised_error flag outside the loop

    # Temporary list to hold libraries that are still needed (those that failed installation)
    remaining_libraries = []

    # Install each library from the requirements.txt file
    for lib in libraries:
        lib = lib.strip()  # Remove leading/trailing whitespace or newlines
        if lib:  # Only attempt to install if the library name is not empty
            logging.info(f"Attempting to install: {lib}")
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", lib])
                logging.info(f"Successfully installed: {lib}")
            except Exception as e:
                raised_error = True
                logging.error(f"Failed to install {lib}: {e}")
                # If the installation fails, keep the library for re-trying later
                remaining_libraries.append(lib)

    # If any installation failed, overwrite the requirements.txt with remaining libraries
    if raised_error:
        with open("requirements.txt", "w") as f:
            for lib in remaining_libraries:
                f.write(lib + "\n")
        logging.info("Updated requirements.txt with remaining libraries.")
        messagebox.showerror("Failed to install required libraries.\nSee log for details.")
        sys.exit(1)  # Exit the program if any installation failed
    else:
        # If all libraries were successfully installed, clear the requirements.txt
        open("requirements.txt", "w").close()  # Empty the file as all libraries are installed
        logging.info("All libraries installed successfully. requirements.txt is now empty.")

install_requirements()

import tkinter as tk
import tkinter.font as tkFont
from tkinter import filedialog, simpledialog, messagebox, ttk
import openpyxl
from openpyxl.styles import PatternFill
import requests
from datetime import datetime
import shutil
import json
import re
import webbrowser

logging.info("All dependencies installed and working correctly. Starting application.")

# Columns
col_A = 1
col_B = 2
col_C = 3

# Initialize file paths
main_file_path = None
new_file_path = None
main_workbook = None
new_workbook = None
searched_order_number = ""
entered_password = ""
main_password = "wachtwoord"
update_url = "https://raw.githubusercontent.com/likejeppy/Pakbon_Editor/refs/heads/main/Editor/Pakbon_Editor.pyw"
latest_version_url = "https://raw.githubusercontent.com/likejeppy/Pakbon_Editor/refs/heads/main/Editor/latest.json"

# Get the directory of the current script
base_dir = os.path.dirname(os.path.abspath(__file__))

# Config file path relative to the script
config_file = os.path.join(base_dir, "config.json")

# Default configuration
default_config = {
    "main_file_path": None,
    "window_position": (686, 350),
    "selection_window_position": (686, 350)
}

def load_config():
    logging.info("Performing function 'load_config'.")
    config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
    if os.path.exists(config_file):
        logging.info("Config file found, loading it.")
        with open(config_file, "r") as f:
            return json.load(f)
    else:
        logging.warning("Config file does not exist, returning default configuration.")
        return {}

def save_config(config):
    logging.info("Performing function 'save_config'.")
    try:
        # Ensure the path is absolute
        if "main_file_path" in config and config["main_file_path"] is not None:
            config["main_file_path"] = os.path.abspath(config["main_file_path"])
        with open(config_file, "w") as f:
            json.dump(config, f, indent=4)
            logging.info(f"Config file saved successfully at: {config_file}")
    except Exception as e:
        logging.error(f"Error saving config file at {config_file}: {e}")

def set_current_version():
    logging.info("Performing function 'set_current_version'.")
    latest_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "latest.json")

    # Check if the file exists
    if os.path.exists(latest_file):
        logging.info("Latest version config file found, loading it.")
        try:
            # Load the content from the existing file
            with open(latest_file, "r") as f:
                file_content = f.read().strip()  # Read and strip any whitespace

                # If the content is empty, update with default values
                if not file_content:
                    logging.warning(f"{latest_file} is empty, updating it with default values.")
                    default_config = {"version": current_version}
                    with open(latest_file, "w") as f_write:
                        json.dump(default_config, f_write, indent=4)
                    logging.info(f"Updated {latest_file} with default values.")
                    return default_config

                # Try loading the JSON content
                config = json.loads(file_content)
                logging.info("Successfully loaded the latest version file.")

                # Update the version in the loaded config
                config["version"] = current_version

                # Save the updated config back to the file
                with open(latest_file, "w") as f_write:
                    json.dump(config, f_write, indent=4)
                logging.info(f"Updated {latest_file} with current version {current_version}.")
                return config

        except json.JSONDecodeError as e:
            logging.error(f"Error reading JSON from {latest_file}: {e}")
            # If there's an error parsing the file, return the default config
            default_config = {"version": current_version}
            with open(latest_file, "w") as f_write:
                json.dump(default_config, f_write, indent=4)
            logging.info(f"Rewritten {latest_file} with default values due to error.")
            return default_config

    else:
        logging.warning(f"{latest_file} does not exist, creating it with default values.")
        # If the file doesn't exist, create it with the default configuration
        default_config = {"version": current_version}
        with open(latest_file, "w") as f_write:
            json.dump(default_config, f_write, indent=4)
        logging.info(f"Created {latest_file} with default values.")
        return default_config

def fetch_online_version():
    logging.info("Performing function 'fetch_online_version'.")
    try:
        # URL to the latest.json file on GitHub
        online_url = latest_version_url
        response = requests.get(online_url)

        # Check if the response is successful
        if response.status_code == 200:
            # Parse the JSON content
            online_config = response.json()
            logging.info(f"Online version info: {online_config}")
            return online_config
        else:
            logging.error(f"Failed to fetch online version info, status code: {response.status_code}")
            return None
    except requests.exceptions.RequestException as e:
        logging.error(f"Error fetching online version info: {e}")
        return None

# Function to check if an update is needed
def check_for_update():
    logging.info("Performing function 'check_for_update'.")
    try:
        # Fetch the latest version info from the raw URL
        response = requests.get(latest_version_url)
        response.raise_for_status()  # Raise an error for invalid responses
        latest_info = response.json()

        # Check if the 'version' key is in the response
        if "version" in latest_info:
            online_version = latest_info["version"]
            logging.info(f"Current version: {current_version}, File version: {online_version}")

            # Compare versions (this is a simple string comparison)
            #test = 1
            if online_version > current_version:
            #if test == 1:
                response = messagebox.askyesno("Update Beschikbaar",
                                               f"Een update is vereist om dit progamma te blijven gebruiken.\nHuidige versie: {current_version}, nieuwe versie: {online_version}\nWil je nu updaten?")
                if response: # response = yes
                    logging.info("Update available, downloading update.")
                    download_update()
                else:
                    logging.info("Update available, but user declined to update.")
                    exit()
            else:
                logging.info("No update required, current version is the most up-to-date version available.")
        else:
            logging.warning("Version info not found in the response.")
    except requests.exceptions.RequestException as e:
        logging.error(f"Error while checking for updates: {e}")

def download_update():
    logging.info("Performing function 'download_update'.")
    try:
        response = requests.get(update_url)

        if response.status_code == 200:
            # Save the updated file to a temporary location
            temp_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Pakbon_Editor_updated.pyw")

            # Write the content to the new file
            with open(temp_file_path, "wb") as f:
                f.write(response.content)
            logging.info("Update downloaded and saved successfully.")

            # Remove the old script and replace it with the updated version
            logging.info("Removing old script and replacing it with the updated version.")

            old_file_path = os.path.abspath(__file__)  # Get the path of the current running script

            # Remove the old file (if it exists)
            if os.path.exists(old_file_path):
                os.remove(old_file_path)
                logging.info("Old script removed.")

            # Move the new file to replace the old one
            shutil.move(temp_file_path, old_file_path)
            logging.info("Updated script is in place.")

            # Use os.execv to relaunch the updated script
            logging.info("Relaunching the updated version of the script.")
            os.execv(sys.executable, [sys.executable, old_file_path])  # This will replace the current script with the new one
        else:
            logging.error(f"Failed to download update, status code: {response.status_code}")
    except requests.exceptions.RequestException as e:
        logging.error(f"Error while downloading update: {e}")

def load_workbook_with_fallback(file_path):
    logging.info("Performing function 'load_workbook_with_fallback'.")
    try:
        return openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        logging.error(f"Het bestand op '{file_path}' is niet gevonden.")
        messagebox.showerror("Bestand niet gevonden", f"Het bestand op '{file_path}' is niet gevonden.")
    except Exception as e:
        logging.error(f"Er heeft een error plaatsgevonden: {e}")
        messagebox.showerror("Error", f"Er heeft een error plaatsgevonden: {e}")
    return None

def password_main_file():
    logging.info("Performing function 'password_main_file'.")
    """Show a dialog to enter the password to browse the main file."""
    global entered_password
    if entered_password is None:
        entered_password = ""
    dialog = EnterPasswordDialog(root, title="Vul Wachtwoord In", initial_value=entered_password)
    if dialog.result:
        return dialog.result
    return None

def browse_main_file():
    logging.info("Performing function 'browse_main_file'.")
    global main_password, entered_password
    while True:
        entered_password = password_main_file()
        if entered_password == main_password:
            global main_file_path, main_workbook
            main_file_path = filedialog.askopenfilename(title="Selecteer Hoofdbestand",
                                                        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
            if main_file_path:
                main_workbook = load_workbook_with_fallback(main_file_path)
                if main_workbook:
                    config = load_config()
                    config['main_file_path'] = main_file_path  # Save the selected file path
                    save_config(config)
                    main_file_button.config(text="Hoofdbestand Geladen", bg="lightgreen")
                    show_file_button.config(bg="lightblue")
                    search_order_button.config(bg="lightblue")
                    new_file_button.config(text="Controleer Nieuwe Pakbon", bg="lightblue")
                    set_latest_date()
            logging.info("Successfully performed function 'browse_main_file'.")
            return
        elif not entered_password is None:
            logging.info("Password is wrong, prompting user again.")
            messagebox.showerror("Error", f"Verkeerd wachtwoord opgegeven: {entered_password}.")
        else:
            return

def reload_main_workbook():
    logging.info("Performing function 'reload_main_workbook'.")
    global main_workbook
    if main_file_path:
        try:
            main_workbook = load_workbook_with_fallback(main_file_path)
            if main_workbook:
                logging.info(f"Successfully reloaded workbook: {main_file_path}")
                # messagebox.showinfo("Success", "Het hoofdbestand is opnieuw geladen.")
                set_latest_date()
            else:
                logging.error(f"Failed to reload workbook: {main_file_path}")
                # messagebox.showerror("Error", "Kan het hoofdbestand niet opnieuw laden.")
        except Exception as e:
            logging.error(f"Error reloading workbook: {e}")
            # messagebox.showerror("Error", f"Fout bij het herladen van het bestand: {e}")
    else:
        logging.warning("No file path set. Cannot reload workbook.")
        # messagebox.showwarning("Waarschuwing", "Geen hoofdbestand ingesteld om opnieuw te laden.")

def browse_new_file():
    logging.info("Performing function 'browse_new_file'.")
    global new_file_path, new_workbook
    new_file_path = filedialog.askopenfilename(title="Selecteer Nieuwe Pakbon",
                                               filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
    if new_file_path:
        new_workbook = load_workbook_with_fallback(new_file_path)
        if new_workbook:
            new_file_button.config(text="Nieuwe Pakbon Geladen", bg="lightgreen")
            logging.info("Successfully performed function 'browse_new_file'.")
            return True
    logging.info("Successfully performed function 'browse_new_file'.")
    return False

def remove_time_if_datetime(value):
    #logging.info("Performing remove_time_if_datetime.") #removed to decrease log spam
    return value.date() if isinstance(value, datetime) else value

def read_column_data(sheet, column):
    logging.info("Performing function 'read_column_data'.")
    return [remove_time_if_datetime(row[0]) for row in
            sheet.iter_rows(min_col=column, max_col=column, values_only=True)]

def style_first_row(sheet):
    logging.info("Performing function 'style_first_row'.")
    # Remove styles and make the first row bold
    for cell in sheet[1]:  # 1 refers to the first row (0-indexed internally)
        cell.fill = openpyxl.styles.PatternFill(start_color="FFFFFF", end_color="FFFFFF",
                                                fill_type="solid")  # Clear fill color
        cell.font = openpyxl.styles.Font(bold=True)
    logging.info("Successfully performed function 'style_first_row'.")

def check_cell_color(cell):
    logging.info("Performing function 'check_cell_color'.")
    """
    Check the background color of the given cell.

    Parameters:
        cell: The openpyxl cell object to check.

    Returns:
        str: Returns "green" if the cell is green (90EE90),
             "red" if the cell is red (FFC0C0),
             or "none" if the cell has any other color or no fill.
    """
    # OpenPyxl prepends 'FF' to RGB color codes
    logging.info("Performing function 'check_cell_color'.")
    green_rgb_00 = "0090EE90"  # Green RGB with '00' prepended
    green_rgb_ff = "FF90EE90"  # Green RGB with 'ff' prepended
    red_rgb_00 = "00FFC0C0"  # Red RGB with '00' prepended
    red_rgb_ff = "FFFFC0C0"  # Red RGB with 'FF' prepended

    if hasattr(cell.fill, 'start_color') and hasattr(cell.fill, 'end_color'):
        start_color = getattr(cell.fill.start_color, 'rgb', None)
        end_color = getattr(cell.fill.end_color, 'rgb', None)

        logging.info(f"Cell colors - Start: {start_color}, End: {end_color}")

        # Check for green
        if start_color == green_rgb_00 or start_color == green_rgb_ff or end_color == green_rgb_00 or end_color == green_rgb_ff:
            logging.info("The cell has a green background.")
            return "green"
        # Check for red
        elif start_color == red_rgb_00 or start_color == red_rgb_ff or end_color == red_rgb_00 or end_color == red_rgb_ff:
            logging.info("The cell has a red background.")
            return "red"

    logging.info("The cell does not have a green or red background.")
    return "none"

def add_data():
    logging.info("Performing function 'add_data'.")
    is_file_selected = browse_new_file()
    if not is_file_selected:
        return
    # messagebox.showinfo("test", str(test_bool))
    if not main_workbook or not new_workbook:
        messagebox.showerror("Error", "Zowel het hoofdbestand als de nieuwe pakbon moeten zijn geladen.")
        return

    # Gather data
    new_sheet = new_workbook.active
    column_a_data = read_column_data(new_sheet, col_A)
    column_b_data = read_column_data(new_sheet, col_B)
    column_c_data = read_column_data(new_sheet, col_C)
    new_sheet_data = list(zip(column_a_data, column_b_data, column_c_data))
    logging.info("Successfylly gathered data in function add_data.")

    # Determine sheet name
    sheet_name = new_sheet["A2"].value
    if isinstance(sheet_name, datetime):
        sheet_name = sheet_name.strftime("%d-%m-%Y")
    sheet_name = str(sheet_name).replace("/", "-")

    if sheet_name in main_workbook.sheetnames:
        response = messagebox.askyesno("Waarschuwing", f"Pakbon van {sheet_name}\nis al eerder verwerkt, toch doorgaan?")
        if not response:  # response = no
            new_file_button.config(text="Controleer Nieuwe Pakbon", bg="lightblue")
            return

    # Prompt user to confirm orders
    selected_indices = confirm_orders(new_sheet_data[1:])
    if selected_indices == []:
        response = messagebox.askyesno("Waarschuwing", "Geen ordernummer afgevinkt, toch toevoegen?")
        if not response:  # response = no
            new_file_button.config(text="Controleer Nieuwe Pakbon", bg="lightblue")
            return
    sheet_date = sheet_name
    # messagebox.showerror("test", f"Selected indices: {selected_indices}")  # Debugging selected indices
    if sheet_name in main_workbook.sheetnames:
        response = messagebox.askyesnocancel("Waarschuwing",
                                             f"Oude data vervangen?")
        if response == True:
            new_file_button.config(text="Controleer Nieuwe Pakbon", bg="lightblue")
            del main_workbook[sheet_name]
        elif response == False:
            # Handle duplicate sheet names
            original_sheet_name = sheet_name
            counter = 1
            while sheet_name in main_workbook.sheetnames:
                match = re.match(r"^(.*)\s\((\d+)\)$", sheet_name)
                if match:
                    sheet_name = f"{match.group(1)} ({int(match.group(2)) + 1})"
                else:
                    sheet_name = f"{original_sheet_name} ({counter})"
                    counter += 1
            new_file_button.config(text="Controleer Nieuwe Pakbon", bg="lightblue")
        else:
            new_file_button.config(text="Controleer Nieuwe Pakbon", bg="lightblue")
            return

    # Create the new sheet
    logging.info("Creating new sheet in function 'add_data'.")
    new_sheet_obj = main_workbook.create_sheet(title=sheet_name)

    # Add data to the sheet and apply colors
    column_widths = [0] * 3  # Track max width for each column
    for row_idx, row in enumerate(new_sheet_data, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = new_sheet_obj.cell(row=row_idx, column=col_idx, value=value)

            # Debugging row_idx and selection status
            # messagebox.showerror("test", f"Processing row {row_idx}, Selected: {row_idx - 1 in selected_indices}")

            # Highlight selected rows in green, others in red
            if (row_idx) - 2 in selected_indices:  # Adjust for zero-based index
                cell.fill = openpyxl.styles.PatternFill(start_color="90EE90", end_color="90EE90",
                                                        fill_type="solid")  # Light green
            else:
                cell.fill = openpyxl.styles.PatternFill(start_color="FFC0C0", end_color="FFC0C0",
                                                        fill_type="solid")  # Light red

            # Adjust column widths
            if value:
                column_widths[col_idx - 1] = max(column_widths[col_idx - 1], len(str(value)))

    style_first_row(new_sheet_obj)

    # Adjust column widths
    for col_idx, width in enumerate(column_widths, start=1):
        new_sheet_obj.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width + 3  # Add padding

    # Save changes
    main_workbook.save(main_file_path)
    logging.info("Successfully performed function 'add_data'.")
    messagebox.showinfo("Success", f"Pakbon van '{sheet_name}' succesvol verwerkt!")
    new_file_button.config(text="Controleer Nieuwe Pakbon", bg="lightblue")
    reload_main_workbook()

class SearchOrderDialog(simpledialog.Dialog):
    logging.info("Setting class 'SearchOrderDialog'.")
    def __init__(self, parent, title=None, initial_value=""):
        self.initial_value = initial_value
        super().__init__(parent, title)
    def body(self, master):
        """Create dialog body."""
        self.title("Bestelling Opzoeken")
        self.resizable(False, False)

        # Centered label
        self.label = tk.Label(self, text="Vul het ordernummer in:", font=("Arial", 12))
        self.label.pack(pady=(20, 5), padx=20)

        # Centered entry field
        self.entry = ttk.Entry(self, font=("Arial", 12), justify="center")
        self.entry.pack(pady=(0, 20), padx=20, ipadx=30)
        self.entry.insert(0, self.initial_value)  # Set initial value

        return self.entry  # Focus on the entry field
    def apply(self):
        """Handle when the dialog is accepted."""
        self.result = self.entry.get()


class EnterPasswordDialog(simpledialog.Dialog):
    logging.info("Setting class 'EnterPasswordDialog'.")

    def __init__(self, parent, title=None, initial_value=""):
        self.initial_value = initial_value
        self.show_password = tk.BooleanVar(value=False)  # Variable to track checkbox state
        super().__init__(parent, title)

    def body(self, master):
        """Create dialog body."""
        self.title("Vul Wachtwoord In")
        self.resizable(False, False)

        # Centered label
        self.label = tk.Label(self, text="Vul het wachtwoord in:", font=("Arial", 12))
        self.label.pack(pady=(20, 5), padx=20)

        # Centered entry field
        self.entry = ttk.Entry(self, font=("Arial", 12), justify="center", show="*")
        self.entry.pack(pady=(0, 5), padx=20, ipadx=30)
        self.entry.insert(0, self.initial_value)  # Set initial value

        # Show password checkbox
        self.show_password_checkbox = ttk.Checkbutton(
            self,
            text="Toon wachtwoord",
            variable=self.show_password,
            command=self.toggle_password_visibility
        )
        self.show_password_checkbox.pack(pady=(0, 0))

        return self.entry  # Focus on the entry field

    def toggle_password_visibility(self):
        """Toggle password visibility and return focus to entry."""
        if self.show_password.get():
            self.entry.config(show="")  # Show password
        else:
            self.entry.config(show="*")  # Hide password
        self.entry.focus_set()  # Return focus to the entry field

    def apply(self):
        """Handle when the dialog is accepted."""
        self.result = self.entry.get()

def search_order_dialog():
    logging.info("Performing function 'search_order_dialog'.")
    """Show a dialog to search for an order number."""
    global searched_order_number
    dialog = SearchOrderDialog(root, title="Bestelling Opzoeken", initial_value=searched_order_number)
    if dialog.result:
        return dialog.result
    return None

def search_order():
    logging.info("Performing function 'search_order'.")
    config = load_config()
    saved_main_file_path = config.get('main_file_path', None)

    while not saved_main_file_path:
        # Show error message
        logging.error("Bestand niet correct geladen in function 'search_order'.")
        response = messagebox.askyesno("Error", "Bestand niet correct geladen! Wil je een bestand selecteren?")

        if response:  # If user clicks 'Yes', open the file dialog
            browse_main_file()  # This function should update the config file
            config = load_config()  # Reload the config after file selection
            saved_main_file_path = config.get('main_file_path', None)  # Re-check the path
        else:  # If user clicks 'No', break out of the loop
            main_file_button.config(text="Selecteer Hoofdbestand", bg="lightgray")
            show_file_button.config(bg="lightgray")
            search_order_button.config(bg="lightgray")
            return

    global searched_order_number
    searched_order_number = ""
    while True:
        logging.info("Performing function 'search_order'.")
        if not main_workbook:
            messagebox.showerror("Error", "Het hoofdbestand moet zijn geladen om de orderstatus te controleren.")
            return

        order_number = search_order_dialog()

        order_number = order_number.strip().lower()
        searched_order_number = order_number
        results = []
        for sheet_name in main_workbook.sheetnames:
            sheet = main_workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value).strip().lower() == order_number:
                        # Find the cell to the left of the current cell
                        left_cell = sheet.cell(row=cell.row, column=cell.column - 1)
                        try:
                            result = check_cell_color(cell)
                            if result == "green":
                                if left_cell.value:
                                    if isinstance(left_cell.value, datetime):
                                        delivery_date = left_cell.value.strftime("%d-%m-%Y")
                                        results.append(f"Bestelling geleverd op: {delivery_date}")
                                    else:
                                        results.append(f"Bestelling geleverd op: {left_cell.value}")
                                else:
                                    results.append("Bestelling geleverd, maar kon geen datum koppelen.")
                                print("The cell is green.")
                            elif result == "red":
                                if left_cell.value:
                                    if isinstance(left_cell.value, datetime):
                                        delivery_date = left_cell.value.strftime("%d-%m-%Y")
                                        results.append(
                                            f"Bestelling hoort bezorgd te zijn op: {delivery_date},\nmaar is niet afgevinkt.")
                                    else:
                                        results.append(
                                            f"Bestelling hoort bezorgd te zijn op: {left_cell.value},\nmaar is niet afgevinkt.")
                                else:
                                    results.append(
                                        "Bestelling gevonden, deze had bezorgd moeten zijn,\nmaar de datum is onbekend en order is niet afgevinkt.")
                                print("The cell is red.")
                            else:
                                if left_cell.value:
                                    if isinstance(left_cell.value, datetime):
                                        delivery_date = left_cell.value.strftime("%d-%m-%Y")
                                        results.append(
                                            f"Bestelling gevonden op pakbon: {delivery_date}.\nLeverstatus is onbekend.")
                                    else:
                                        results.append(
                                            f"Bestelling gevonden op pakbon: {left_cell.value}.\nLeverstatus is onbekend.")
                                else:
                                    results.append(
                                        "Bestelling gevonden, maar kon geen datum koppelen.\nLeverstatus is onbekend.")
                                print("The cell does not have a red or green background.")
                            break
                        except Exception as e:
                            logging.error(
                                f"Error in search_order after searching for order: {searched_order_number}. Error: {str(e)}")
                            messagebox.showerror("Error in search_orders", str(e))
        if results:
            logging.info(f"Searched for order: {searched_order_number}, order was successfully found.")
            messagebox.showinfo("Bestelling Gevonden", "\n".join(results))
            #return #enable to auto close search panel after successfully found onrder
        else:
            logging.info(f"Searched for order: {searched_order_number}, but order not found.")
            messagebox.showerror("Bestelling Niet Gevonden", f"Het ordernummer: {searched_order_number}\nis niet gevonden.")

def confirm_orders(data):
    logging.info("Performing function 'confirm_orders'.")
    """
    Displays a window for the user to select rows from the given data.
    Returns a list of indices of selected rows.
    """
    if not data:
        messagebox.showerror("Error", "Geen gegevens gevonden in de pakbon.")
        return []

    selected_indices = []

    def on_confirm():
        logging.info("Performing function 'on_confirm' in function 'confirm_orders'.")
        for idx, var in enumerate(checkbox_vars):
            if var.get():
                selected_indices.append(original_indices[idx])  # Map back to original index
        selection_window.destroy()

    # Create a new window for order selection
    logging.info("Creating window in function 'confirm_orders'.")
    selection_window = tk.Toplevel(root)
    selection_window.title("Bevestig Orders")
    selection_window.resizable(False, False)
    selection_window_position = config.get('selection_window_position',
                                           (root.winfo_x(), root.winfo_y()))  # Default to (100, 100)
    selection_window.geometry(f"+{selection_window_position[0]}+{selection_window_position[1]}")
    selection_window.resizable(False, False)

    # Calculate the maximum width needed
    font = tkFont.Font(family="Arial", size=12)
    checkbox_texts = [f"{row[1]} Selecteer" for row in data]
    max_text_width = max(font.measure(text) for text in checkbox_texts)
    max_width = max(200, max_text_width + 50)  # Add padding change 200 to larger to increase width

    # Label at the top
    tk.Label(selection_window, text="Selecteer de orders die\nbevestigd moeten worden", font=("Arial", 13),
             wraplength=max_width).pack(pady=10)

    # Create a frame for the scrollable area
    scrollable_frame = tk.Frame(selection_window)
    scrollable_frame.pack(fill=tk.BOTH, expand=True)

    # Create a canvas and a scrollbar for the frame
    canvas = tk.Canvas(scrollable_frame, width=max_width)
    scrollbar = tk.Scrollbar(scrollable_frame, orient="vertical", command=canvas.yview)
    scrollable_content = tk.Frame(canvas)
    scrollable_content.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )

    canvas.create_window((0, 0), window=scrollable_content, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    # Pack the canvas and scrollbar
    canvas.pack(side="left", fill=tk.BOTH, expand=True)
    scrollbar.pack(side="right", fill="y")

    min_to_scroll = 9
    if len(data) >= min_to_scroll:
        # Enable scrolling with the mouse wheel
        def on_mouse_wheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", on_mouse_wheel)

    # Pair original index with each row BEFORE sorting
    indexed_data = list(enumerate(data))
    # Sort by last 4 characters of the order number (row[1][1] is the order number)
    sorted_data = sorted(indexed_data, key=lambda row: row[1][1][-4:])
    # Split sorted data into separate lists
    original_indices = [idx for idx, row in sorted_data]
    sorted_rows = [row for idx, row in sorted_data]
    # messagebox.showinfo("test", original_indices) # have to be unique to handle duplicates!
    checkbox_vars = []
    for idx, row in enumerate(sorted_rows):
        order_number = row[1]  # Assuming row is a tuple like (col_A, col_B, col_C), and order number is in col_B

        # Separate the last four characters and make them bold
        base_text = order_number[:-4]  # All but the last 4 characters
        bold_text = order_number[-4:]  # The last 4 characters to be bold

        var = tk.BooleanVar()
        checkbox_vars.append(var)

        # Create a frame for each order number to hold the labels
        order_frame = tk.Frame(scrollable_content)
        order_frame.pack(anchor="w", pady=2)

        # Display the non-bold part of the order number
        label_base = tk.Label(order_frame, text=base_text, font=("Arial", 12), relief="flat", anchor="w",
                              wraplength=max_width - 100)
        label_base.pack(side="left", padx=0, pady=0)

        # Display the bold part of the order number
        label_bold = tk.Label(order_frame, text=bold_text, font=("Arial", 12, "bold"), relief="flat", anchor="w")
        label_bold.pack(side="left", padx=0, pady=0)

        # Add a Checkbutton for selecting the order
        checkbutton = tk.Checkbutton(order_frame, text="Selecteer", variable=var, font=("Arial", 12))
        checkbutton.pack(side="left", padx=10)

    # Confirm button to submit selections
    tk.Button(selection_window, text="Bevestig", font=("Arial", 12), command=on_confirm, bg="lightblue").pack(pady=10,
                                                                                                              fill=tk.X)
    # Declare variables for Select/Deselect button functionality
    AllSelected = [False]  # Use a list to make it mutable
    SelectButtonText = tk.StringVar()
    SelectButtonText.set("Selecteer Alles")  # Set initial button text

    # Select All/Deselect All button for convenience
    def select_all():
        if AllSelected[0] == False:
            for var in checkbox_vars:
                var.set(True)  # Select all checkboxes
            AllSelected[0] = True
            SelectButtonText.set("Deselecteer Alles")  # Change button text to Deselecteer Alles
        else:
            for var in checkbox_vars:
                var.set(False)  # Deselect all checkboxes
            AllSelected[0] = False
            SelectButtonText.set("Selecteer Alles")  # Change button text back to Selecteer Alles

    # Create Select All/Deselect All button with dynamic text and stretch it across the window
    SelectButton = tk.Button(selection_window, textvariable=SelectButtonText, command=select_all, font=("Arial", 12),
                             bg="lightblue")
    SelectButton.pack(pady=5, fill="x")

    # Wait for the user to close the window
    selection_window.grab_set()
    selection_window_position = (root.winfo_x(), root.winfo_y())  # to use main window location
    config['selection_window_position'] = selection_window_position
    root.wait_window(selection_window)

    # Return the list of selected row indices
    logging.info("Successfully performed function 'confirm_orders'.")
    return selected_indices

def testfunc():
    logging.info("Performing function 'testfunc'.")
    global main_workbook
    messagebox.showinfo("Workbook", f"{main_workbook}")

def test_confirm_orders():
    logging.info("Performing function 'test_confirm_orders'.")
    # Mock dataset to simulate data from columns A, B, and C
    mock_data = [
        ("Order1", "ProductA", "CustomerX"),
        ("Order2", "ProductB", "CustomerY"),
        ("Order3", "ProductC", "CustomerZ"),
    ]

    # Call the confirm_orders function with the mock data
    selected_orders = confirm_orders(mock_data)

    # Display the selected rows in a messagebox
    if selected_orders:
        messagebox.showinfo("Geselecteerde Orders", f"Je hebt deze orders geselecteerd:\n{selected_orders}")
    else:
        messagebox.showinfo("Geen Selectie", "Er zijn geen orders geselecteerd.")


def open_and_display_excel_file():
    logging.info("Performing function 'open_and_display_excel_file'.")
    config = load_config()
    saved_main_file_path = config.get('main_file_path', None)

    while not saved_main_file_path:
        # Show error message
        response = messagebox.askyesno("Error", "Bestand niet correct geladen! Wil je een bestand selecteren?")

        if response:  # If user clicks 'Yes', open the file dialog
            browse_main_file()  # This function should update the config file
            config = load_config()  # Reload the config after file selection
            saved_main_file_path = config.get('main_file_path', None)  # Re-check the path
            return
        else:  # If user clicks 'No', break out of the loop
            main_file_button.config(text="Selecteer Hoofdbestand", bg="lightgray")
            show_file_button.config(bg="lightgray")
            search_order_button.config(bg="lightgray")
            return
    test_saved_main_file_path = config.get('main_file_path', None)

    if test_saved_main_file_path:
        test_main_file_path = saved_main_file_path
        test_main_workbook = load_workbook_with_fallback(main_file_path)
        if test_main_workbook:
            # file is loaded correctly
            logging.info("File loaded correctly in function 'open_and_display_excel_file'.")
            file_url = "https://hemaoffice-my.sharepoint.com/:x:/r/personal/fm-0337_hema_nl/_layouts/15/Doc.aspx?sourcedoc=%7B7351D898-572E-4BC2-BBD0-9FAC1FF1AAFC%7D&file=Pakbon.xlsx&action=default&mobileredirect=true"  # Replace with actual file URL, try to sync with cloud?
            open_in_excel_online(file_url)
            # messagebox.showerror("Error", "Deze functie bestaat nog niet.")
        else:
            # file not loaded correctly
            logging.info("Failed to load file in function 'open_and_display_excel_file'.")
            messagebox.showerror("Error", "Bestand is niet correct geladen!")
    else:
        logging.error("Error in function 'open_and_display_excel_file'.")
        messagebox.showerror("Error", "Bestand is niet correct geladen!")

def open_in_excel_online(file_url):
    logging.info("Performing function 'open_in_excel_online'.")
    edge_path = "C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
    webbrowser.register('edge', None, webbrowser.BackgroundBrowser(edge_path))
    webbrowser.get('edge').open(f"{file_url}")
    logging.info("Successfully performed function 'open_in_excel_online'.")

def clear_logs_file(log_file_path="app.log"):
    logging.info("Performing function 'clear_logs_file'.")
    """
    Clears all logs in the specified log file.

    Parameters:
        log_file_path (str): The path to the log file..
    """
    response = messagebox.askyesno("Waarschuwing", f"Weet je zeker dat je alle logs wilt verwijderen?\nDeze actie kan niet ongedaan worden.")
    if response:  # response = yes
        try:
            with open(log_file_path, 'w') as log_file:
                log_file.write("")  # Overwrite the file with an empty string
            messagebox.showinfo("Succes", "Logs met succes verwijderd!")
        except Exception as e:
            messagebox.showerror(f"Error clearing logs: {e}")
    return

def set_latest_date():
    logging.info("Performing function 'set_latest_date'.")
    global main_workbook
    if main_file_path:
        try:
            # Reload the workbook
            main_workbook = load_workbook_with_fallback(main_file_path)
            if main_workbook:
                # Collect all sheet names
                sheet_names = main_workbook.sheetnames

                # Parse sheet names as dates and find the latest date
                # List of supported date formats
                date_formats = ["%d-%m-%Y", "%m-%d-%Y"]
                latest_date = None
                for name in sheet_names:
                    for date_format in date_formats:
                        try:
                            # Attempt to parse the sheet name as a date in the current format
                            date = datetime.strptime(name, date_format)
                            if latest_date is None or date > latest_date:
                                latest_date = date
                            break  # Exit the loop if parsing succeeds
                        except ValueError:
                            continue  # Try the next format if the current one fails
                    else:
                        # Log a warning if none of the formats matched
                        logging.warning(f"Sheet name '{name}' is not a valid date format in any of {date_formats}")

                if latest_date:
                    print(f"Latest date: {latest_date.strftime('%d-%m-%Y')}")
                else:
                    print("No valid dates found.")

                # Update the label with the latest date
                if latest_date:
                    latest_date_str = latest_date.strftime("%d-%m-%Y")
                    lastloaded.config(text=f"Meest recente levering:\n{latest_date_str}")
                    logging.info(f"Set label text to: {latest_date_str}")
                else:
                    lastloaded.config(text="Meest recente levering:\nGeen eerdere data gevonden")
                    logging.info("Geen eerdere data gevonden.")
            else:
                lastloaded.config(text="Meest recente levering:\nGeen eerdere data gevonden")
                logging.info("Geen eerdere data gevonden (couldn't load main workbook.")
        except Exception as e:
            logging.error(f"Error while loading workbook or getting value: {e}")
            messagebox.showerror("Error", f"Error: {e}")

# GUI setup
root = tk.Tk()
root.title(f"Pakbon Editor - jeffvh {current_version}")

logging.info("Loading config file.")
config = load_config()

frame = tk.Frame(root, padx=20, pady=10)
frame.pack()

tk.Label(frame, text="Pakbon Editor", font=("Arial", 16)).grid(row=0, column=0, columnspan=2, pady=(0, 0))

lastloaded = tk.Label(frame, text="", font=("Arial", 10))
lastloaded.grid(row=1, column=0, columnspan=2, pady=(0, 0))

# Check if there's a saved main file path
saved_main_file_path = config.get('main_file_path', None)

if saved_main_file_path:
    main_file_path = saved_main_file_path
    # Assume load_workbook_with_fallback is defined elsewhere
    main_workbook = load_workbook_with_fallback(main_file_path)
    if main_workbook:
        main_file_button_text = "Hoofdbestand Geladen"
        main_file_button_color = "lightgreen"
        show_file_button_color = "lightblue"
        search_order_button_color = "lightblue"
        new_file_button_color = "lightblue"
    else:
        main_file_button_text = "Selecteer Hoofdbestand"
        main_file_button_color = "lightgray"
        show_file_button_color = "lightgray"
        search_order_button_color = "lightgray"
        new_file_button_color = "lightgray"
else:
    main_file_button_text = "Selecteer Hoofdbestand"
    main_file_button_color = "lightgray"
    show_file_button_color = "lightgray"
    search_order_button_color = "lightgray"
    new_file_button_color = "lightgray"

# Set window position if it exists in the config
window_position = config.get('window_position', (100, 100))  # Default to (100, 100)
root.geometry(f"+{window_position[0]}+{window_position[1]}")
root.resizable(False, False)

# Test confirm_orders button
# tk.Button(frame, text="Test Order Selectie", command=test_confirm_orders, width=20, bg="lightblue").grid(row=3, column=0, pady=5)

# Main file browse button
main_file_button = tk.Button(frame, text=main_file_button_text, command=browse_main_file, width=20,
                             bg=main_file_button_color)
main_file_button.grid(row=2, column=0, pady=5)

# New file browse button
new_file_button = tk.Button(frame, text="Controleer Nieuwe Pakbon", command=add_data, width=20,
                            bg=new_file_button_color)
new_file_button.grid(row=2, column=1, pady=5)

# Search order button
search_order_button = tk.Button(frame, text="Zoek Ordernummer", command=search_order, width=20,
                                bg=search_order_button_color)
search_order_button.grid(row=3, column=0, pady=5)

# Add data button
# tk.Button(frame, text="Controleer Nieuwe Pakbon", command=add_data, width=20, bg="lightblue").grid(row=2, column=1, pady=5)

# Open main file online button
show_file_button = tk.Button(frame, text="Open Bestand Online", command=open_and_display_excel_file, width=20,
                             bg=show_file_button_color)
show_file_button.grid(row=4, column=0, pady=5)

tk.Button(frame, text="Verwijder Logs", command=clear_logs_file, width=20, bg="lightblue").grid(row=3, column=1,
                                                                                                pady=5)  # clear log button, debug only?

# tk.Button(frame, text="Test", command=testfunc, width=20, bg="lightblue").grid(row=3, column=1, pady=5) # test button

# Save window position and file path on close
def on_close():
    logging.info("Performing function 'on_close'.")
    window_position = (root.winfo_x(), root.winfo_y())
    config['window_position'] = window_position
    config['main_file_path'] = main_file_path  # Save the main file path
    save_config(config)
    logging.info("Successfully performed function 'on_close'.")
    logging.info(
        "Closing application...")  # Last log entry
    root.destroy()

root.protocol("WM_DELETE_WINDOW", on_close)

set_latest_date()
set_current_version()
check_for_update()

root.mainloop()
